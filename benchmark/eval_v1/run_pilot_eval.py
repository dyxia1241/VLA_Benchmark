#!/usr/bin/env python3
"""
Pilot evaluation script for GM-100 benchmark via OpenAI-compatible API.

Designed for DashScope compatible endpoint (Qwen VL), but works for any
OpenAI-compatible chat-completions endpoint that supports image_url inputs.
"""

from __future__ import annotations

import argparse
import asyncio
import base64
import hashlib
import itertools
import json
import os
import random
import re
import time
from collections import OrderedDict
from io import BytesIO
from pathlib import Path
from typing import Any

import aiohttp
from PIL import Image, ImageDraw, ImageFont, ImageOps
from tqdm.asyncio import tqdm_asyncio


# Centralized hyperparameter/config entry points.
# You can edit these defaults directly in-file, and still override via CLI args.
BENCHMARK_DIR = Path(__file__).resolve().parent.parent
REPO_ROOT = BENCHMARK_DIR.parent
DATASET_ROOT = REPO_ROOT / "gm100-cobotmagic-lerobot"
PILOT_ARTIFACT_DIR = BENCHMARK_DIR / "previous_results" / "manual_checks_20260319"

DEFAULT_INPUT_JSONL = str(PILOT_ARTIFACT_DIR / "pilot_qa_raw_filtered.jsonl")
DEFAULT_OUTPUT_JSONL = str(BENCHMARK_DIR / "eval_results_v1" / "pilot_results_qwen3vl.jsonl")
DEFAULT_TASK_META_XLSX = str(BENCHMARK_DIR / "GM100 List.xlsx")

DEFAULT_API_BASE = "http://35.220.164.252:3888/v1/"
DEFAULT_MODEL = "gpt-4o"
# Leave empty by default. Pass --api-key or set provider env vars.
DEFAULT_API_KEY = "sk-LIBwaAEBArGfoun34xCXZqDaPYog9imwTx3ZTN1u88p5fFY9"

DEFAULT_CONCURRENCY = 8
DEFAULT_MAX_TOKENS = 96
DEFAULT_TEMPERATURE = 0.0
DEFAULT_RETRIES = 3
DEFAULT_TIMEOUT_SEC = 60.0
DEFAULT_LIMIT = 0  # 0 = all items

DEFAULT_FRAME_DIR = str(PILOT_ARTIFACT_DIR / "pilot_frames_filtered")
DEFAULT_FRAME_DIR_T3 = str(PILOT_ARTIFACT_DIR / "pilot_frames_t3ctx2_filtered")
DEFAULT_FRAME_DIR_T6 = str(PILOT_ARTIFACT_DIR / "pilot_frames_t6_5ctx_filtered")
DEFAULT_T3_OFFSETS = "-10,-5,0,5"
GM100_FALLBACK_CACHE_DIR = BENCHMARK_DIR / "eval_results_v1" / "_frame_fallback_cache" / "gm100"
GM100_VIDEO_READER_CACHE_SIZE = 2

T4_CHOICES = {
    "A": "both arms are active",
    "B": "only the left arm is active",
    "C": "only the right arm is active",
    "D": "both arms are idle",
}
T4_LABEL_ID_TO_ANSWER = {0: "A", 1: "B", 2: "C", 3: "D"}
SLUG_RE = re.compile(r"^[a-z0-9]+(?:-[a-z0-9]+)+$")
ANSWER_TAG_RE = re.compile(r"<\s*ANSWER\s*>(.*?)<\s*/\s*ANSWER\s*>", flags=re.IGNORECASE | re.DOTALL)
_GM100_VIDEO_READERS: OrderedDict[str, Any] = OrderedDict()


def img_to_b64(path: Path) -> str:
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def pil_to_b64(img: Image.Image) -> str:
    buf = BytesIO()
    img.save(buf, format="JPEG", quality=95)
    return base64.b64encode(buf.getvalue()).decode("utf-8")


try:
    _RESAMPLE = Image.Resampling.LANCZOS
except AttributeError:
    _RESAMPLE = Image.LANCZOS


_BINARY_PANEL_SIZE = (640, 640)
_BINARY_PANEL_MARGIN = 24
_BINARY_PANEL_GUTTER = 28
_BINARY_LABEL_BAR_H = 78
_BINARY_FONT_CANDIDATES = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
]


def _load_binary_label_font(size: int) -> ImageFont.ImageFont:
    for cand in _BINARY_FONT_CANDIDATES:
        try:
            p = Path(cand)
            if p.exists():
                return ImageFont.truetype(str(p), size=size)
        except Exception:
            continue
    try:
        return ImageFont.truetype("DejaVuSans-Bold.ttf", size=size)
    except Exception:
        return ImageFont.load_default()


def _measure_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> tuple[int, int]:
    if hasattr(draw, "textbbox"):
        left, top, right, bottom = draw.textbbox((0, 0), text, font=font)
        return right - left, bottom - top
    return draw.textsize(text, font=font)


def _t_binary_display_labels(item: dict[str, Any], num_panels: int = 2) -> list[str]:
    labels = [str(x).upper() for x in item.get("display_labels", []) if str(x).strip()]
    fallback = ["X", "Y"]
    for lbl in fallback:
        if len(labels) >= num_panels:
            break
        if lbl not in labels:
            labels.append(lbl)
    while len(labels) < num_panels:
        labels.append(f"IMG{len(labels) + 1}")
    return labels[:num_panels]


def _t_binary_label_to_answer(item: dict[str, Any]) -> dict[str, str]:
    labels = _t_binary_display_labels(item, num_panels=2)
    choice_keys = [str(k).upper() for k in item.get("choices", {}).keys()]
    choice_set = set(choice_keys)
    label_set = set(labels)

    if choice_set == label_set and len(choice_set) == 2:
        return {label: label for label in sorted(label_set)}
    if choice_set == {"A", "B"}:
        return {labels[0]: "A", labels[1]: "B"}
    if label_set == {"X", "Y"}:
        return {label: label for label in sorted(label_set)}
    return {labels[0]: labels[0], labels[1]: labels[1]}


def _render_t_binary_panel(img: Image.Image, label: str) -> Image.Image:
    panel_w, panel_h = _BINARY_PANEL_SIZE
    outer_bg = (236, 233, 228)
    card_bg = (248, 247, 243)
    ink = (24, 24, 24)
    image_bg = (255, 255, 255)

    panel = Image.new("RGB", (panel_w, panel_h), outer_bg)
    draw = ImageDraw.Draw(panel)
    draw.rounded_rectangle([0, 0, panel_w - 1, panel_h - 1], radius=28, fill=card_bg, outline=ink, width=4)

    bar = [
        _BINARY_PANEL_MARGIN,
        _BINARY_PANEL_MARGIN,
        panel_w - _BINARY_PANEL_MARGIN,
        _BINARY_PANEL_MARGIN + _BINARY_LABEL_BAR_H,
    ]
    draw.rounded_rectangle(bar, radius=18, fill=ink)

    font = _load_binary_label_font(max(28, int(_BINARY_LABEL_BAR_H * 0.58)))
    tw, th = _measure_text(draw, label, font)
    tx = bar[0] + (bar[2] - bar[0] - tw) // 2
    ty = bar[1] + (bar[3] - bar[1] - th) // 2 - 2
    draw.text((tx, ty), label, font=font, fill=(245, 245, 242))

    frame_box = [
        _BINARY_PANEL_MARGIN,
        bar[3] + 16,
        panel_w - _BINARY_PANEL_MARGIN,
        panel_h - _BINARY_PANEL_MARGIN,
    ]
    draw.rounded_rectangle(frame_box, radius=18, fill=image_bg, outline=ink, width=3)

    inner_w = frame_box[2] - frame_box[0] - 24
    inner_h = frame_box[3] - frame_box[1] - 24
    fitted = ImageOps.contain(img, (inner_w, inner_h), method=_RESAMPLE)
    px = frame_box[0] + 12 + (inner_w - fitted.width) // 2
    py = frame_box[1] + 12 + (inner_h - fitted.height) // 2
    panel.paste(fitted, (px, py))
    return panel


def build_t_binary_composite_b64(frame_paths: list[Path], item: dict[str, Any]) -> str:
    labels = _t_binary_display_labels(item, num_panels=len(frame_paths))
    panels: list[Image.Image] = []
    for label, path in zip(labels, frame_paths):
        with Image.open(path) as raw:
            panels.append(_render_t_binary_panel(raw.convert("RGB"), label))

    panel_w, panel_h = _BINARY_PANEL_SIZE
    margin = 24
    gutter = _BINARY_PANEL_GUTTER
    canvas_w = margin * 2 + len(panels) * panel_w + max(0, len(panels) - 1) * gutter
    canvas_h = margin * 2 + panel_h
    canvas = Image.new("RGB", (canvas_w, canvas_h), (230, 227, 221))

    x = margin
    for panel in panels:
        canvas.paste(panel, (x, margin))
        x += panel_w + gutter

    return pil_to_b64(canvas)


def _candidate_single_paths(base: Path, task_id: str, episode_id: int, frame_index: int, camera: str, suffix: str = "") -> list[Path]:
    # Existing files are typically unpadded episode ids, but keep a padded fallback.
    name_unpadded = f"{task_id}_{episode_id}_{frame_index}_{camera}"
    name_padded = f"{task_id}_{episode_id:06d}_{frame_index}_{camera}"
    if suffix:
        return [base / f"{name_unpadded}_{suffix}.jpg", base / f"{name_padded}_{suffix}.jpg"]
    return [base / f"{name_unpadded}.jpg", base / f"{name_padded}.jpg"]


def _candidate_reassemble_single_paths(base: Path, recording_id: str, frame_index: int, camera: str) -> list[Path]:
    return [base / f"{recording_id}_{frame_index}_{camera}.jpg"]


def _uses_recording_id_frame_names(dataset: str) -> bool:
    return dataset in {"REASSEMBLE", "RH20T", "AIST", "AIST-BIMANUAL"}


def _gm100_video_path(task_id: str, episode_id: int, camera: str) -> Path:
    return (
        DATASET_ROOT
        / task_id
        / "videos"
        / "chunk-000"
        / f"observation.images.{camera}"
        / f"episode_{episode_id:06d}.mp4"
    )


def _gm100_video_reader(vpath: Path) -> Any:
    import decord

    decord.bridge.set_bridge("native")
    key = str(vpath)
    vr = _GM100_VIDEO_READERS.pop(key, None)
    if vr is None:
        vr = decord.VideoReader(str(vpath), ctx=decord.cpu(0))
    _GM100_VIDEO_READERS[key] = vr
    while len(_GM100_VIDEO_READERS) > GM100_VIDEO_READER_CACHE_SIZE:
        _GM100_VIDEO_READERS.popitem(last=False)
    return vr


def _extract_gm100_frame(task_id: str, episode_id: int, frame_index: int, camera: str, out_path: Path) -> Path:
    vpath = _gm100_video_path(task_id=task_id, episode_id=episode_id, camera=camera)
    if not vpath.exists():
        raise FileNotFoundError(f"gm100 video missing: {vpath}")
    vr = _gm100_video_reader(vpath)
    if frame_index < 0 or frame_index >= len(vr):
        raise IndexError(f"frame_index out of range: {frame_index} (len={len(vr)}) for {vpath}")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    img = Image.fromarray(vr[frame_index].asnumpy())
    img.save(out_path, format="JPEG", quality=95)
    return out_path


def _clear_gm100_video_reader_cache() -> None:
    _GM100_VIDEO_READERS.clear()


def _requested_gm100_frames(item: dict[str, Any], t3_offsets: list[int]) -> list[tuple[int, int]]:
    ep = item.get("episode_id", None)
    task_type = str(item.get("task_type", ""))

    if task_type == "T2" and "frame_A" in item and "frame_B" in item:
        fa = item["frame_A"]
        fb = item["frame_B"]
        return [
            (int(fa["episode_id"]), int(fa["frame_index"])),
            (int(fb["episode_id"]), int(fb["frame_index"])),
        ]

    if ep is None:
        return []

    ep_int = int(ep)
    if task_type == "T3":
        frame_indices = item.get("frame_indices")
        if isinstance(frame_indices, list) and frame_indices:
            return [(ep_int, int(fi)) for fi in frame_indices]
        if "frame_index" in item:
            center = int(item["frame_index"])
            return [(ep_int, int(center + off)) for off in t3_offsets]
        return []

    if task_type == "T6" and "frame_index" in item:
        center = int(item["frame_index"])
        return [(ep_int, int(center + off)) for off in (-6, -3, 0, 3, 6)]

    frame_indices = item.get("frame_indices")
    if isinstance(frame_indices, list) and frame_indices:
        return [(ep_int, int(fi)) for fi in frame_indices]

    if "frame_index" in item:
        return [(ep_int, int(item["frame_index"]))]

    return []


def resolve_missing_frame_paths(item: dict[str, Any], frame_paths: list[Path], t3_offsets: list[int]) -> list[Path]:
    dataset = normalize_text(item.get("dataset", "")).upper()
    if dataset != "GM100":
        return frame_paths

    task_id = normalize_text(item.get("task_id", ""))
    camera = normalize_text(item.get("camera", "")) or "camera_top"
    requests = _requested_gm100_frames(item, t3_offsets)
    if not task_id or len(requests) != len(frame_paths):
        return frame_paths

    resolved: list[Path] = []
    for (episode_id, frame_index), frame_path in zip(requests, frame_paths):
        if frame_path.exists():
            resolved.append(frame_path)
            continue

        fallback_path = GM100_FALLBACK_CACHE_DIR / frame_path.name
        if fallback_path.exists():
            resolved.append(fallback_path)
            continue

        try:
            resolved.append(
                _extract_gm100_frame(
                    task_id=task_id,
                    episode_id=int(episode_id),
                    frame_index=int(frame_index),
                    camera=camera,
                    out_path=fallback_path,
                )
            )
        except Exception:
            resolved.append(frame_path)

    return resolved


def _gm100_missing_frame_requests(
    item: dict[str, Any],
    frame_paths: list[Path],
    t3_offsets: list[int],
) -> list[tuple[str, int, int, str, Path]]:
    dataset = normalize_text(item.get("dataset", "")).upper()
    if dataset != "GM100":
        return []

    task_id = normalize_text(item.get("task_id", ""))
    camera = normalize_text(item.get("camera", "")) or "camera_top"
    requests = _requested_gm100_frames(item, t3_offsets)
    if not task_id or len(requests) != len(frame_paths):
        return []

    out: list[tuple[str, int, int, str, Path]] = []
    for (episode_id, frame_index), frame_path in zip(requests, frame_paths):
        if frame_path.exists():
            continue
        fallback_path = GM100_FALLBACK_CACHE_DIR / frame_path.name
        if fallback_path.exists():
            continue
        out.append((task_id, int(episode_id), int(frame_index), camera, fallback_path))
    return out


def prewarm_gm100_missing_frames(
    items: list[dict[str, Any]],
    frame_dirs: dict[str, Path],
    t3_offsets: list[int],
) -> dict[str, Any]:
    unique_requests: dict[str, tuple[str, int, int, str, Path]] = {}
    gm100_items = 0
    skipped_items = 0

    for item in items:
        x = normalize_item_for_eval(item)
        if normalize_text(x.get("dataset", "")).upper() != "GM100":
            continue
        gm100_items += 1
        try:
            frame_paths = get_frame_paths(x, frame_dirs=frame_dirs, t3_offsets=t3_offsets)
        except Exception:
            skipped_items += 1
            continue
        for request in _gm100_missing_frame_requests(x, frame_paths, t3_offsets=t3_offsets):
            unique_requests.setdefault(str(request[-1]), request)

    planned = sorted(unique_requests.values(), key=lambda x: (x[0], x[1], x[3], x[2]))
    if not planned:
        return {
            "gm100_items": gm100_items,
            "skipped_items": skipped_items,
            "planned_frames": 0,
            "extracted_frames": 0,
            "failed_frames": 0,
        }

    print(
        f"[INFO] Prewarming {len(planned)} missing GM100 frame(s) into "
        f"{GM100_FALLBACK_CACHE_DIR}"
    )
    extracted = 0
    failed = 0
    for idx, (task_id, episode_id, frame_index, camera, out_path) in enumerate(planned, start=1):
        try:
            _extract_gm100_frame(
                task_id=task_id,
                episode_id=episode_id,
                frame_index=frame_index,
                camera=camera,
                out_path=out_path,
            )
            extracted += 1
        except Exception as e:  # noqa: BLE001
            failed += 1
            if failed <= 5:
                print(
                    "[WARN] GM100 prewarm failed "
                    f"(task={task_id}, episode={episode_id}, frame={frame_index}, camera={camera}): {e}"
                )
        if idx % 200 == 0 or idx == len(planned):
            print(
                f"[INFO] GM100 prewarm progress {idx}/{len(planned)} "
                f"(ok={extracted}, failed={failed})"
            )

    _clear_gm100_video_reader_cache()
    return {
        "gm100_items": gm100_items,
        "skipped_items": skipped_items,
        "planned_frames": len(planned),
        "extracted_frames": extracted,
        "failed_frames": failed,
    }


def _candidate_multi_paths(
    base: Path,
    task_id: str,
    episode_id: int,
    center_frame_index: int,
    camera: str,
    offset: int,
    suffix: str,
) -> list[Path]:
    # Convention A: center-index naming, e.g. ..._159_camera_top_t-2.jpg
    center_paths = _candidate_single_paths(
        base,
        task_id=task_id,
        episode_id=episode_id,
        frame_index=center_frame_index,
        camera=camera,
        suffix=suffix,
    )
    # Convention B: shifted-index naming, e.g. ..._157_camera_top_t-2.jpg
    shifted_paths = _candidate_single_paths(
        base,
        task_id=task_id,
        episode_id=episode_id,
        frame_index=center_frame_index + offset,
        camera=camera,
        suffix=suffix,
    )
    return center_paths + shifted_paths


def _pick_first_existing(paths: list[Path]) -> Path:
    for p in paths:
        if p.exists():
            return p
    # Return first candidate for better missing-path diagnostics.
    return paths[0]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").strip()
    return " ".join(text.split())


def pretty_task_name(task_name: str) -> str:
    return str(task_name or "").replace("-", " ").strip()


def is_slug_like(text: str) -> bool:
    return bool(SLUG_RE.fullmatch(normalize_text(text)))


def load_dataset_task_names() -> dict[str, str]:
    task_names: dict[str, str] = {}
    for meta_path in sorted(DATASET_ROOT.glob("task_*/meta/tasks.jsonl")):
        task_id = meta_path.parts[-3]
        with meta_path.open("r", encoding="utf-8") as f:
            first = next((line for line in f if line.strip()), "")
        if not first:
            continue
        obj = json.loads(first)
        task_name = pretty_task_name(str(obj.get("task", "")).strip())
        if task_name:
            task_names[task_id] = task_name
    return task_names


def load_task_meta_descriptions(xlsx_path: Path) -> dict[str, str]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover - depends on local env
        raise RuntimeError(
            "Task-meta prompting requires openpyxl. Install it or disable --prepend-task-meta."
        ) from e

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Task-meta workbook not found: {xlsx_path}")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "link" not in workbook.sheetnames:
        raise ValueError(f'Workbook {xlsx_path} does not contain sheet "link".')

    descriptions_by_task: dict[str, list[str]] = {}
    sheet = workbook["link"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_task_id = row[0]
        raw_description = row[1]
        if raw_task_id is None:
            continue
        task_id = f"task_{int(raw_task_id):05d}"
        description = normalize_text(raw_description)
        if description:
            descriptions_by_task.setdefault(task_id, []).append(description)

    dataset_fallback = load_dataset_task_names()
    task_ids = sorted(set(descriptions_by_task) | set(dataset_fallback))
    resolved: dict[str, str] = {}
    for task_id in task_ids:
        candidates = list(dict.fromkeys(descriptions_by_task.get(task_id, [])))
        natural = next((x for x in candidates if not is_slug_like(x)), "")
        slug = next((x for x in candidates if is_slug_like(x)), "")
        chosen = natural or pretty_task_name(slug) or dataset_fallback.get(task_id, "")
        if not chosen and task_id in dataset_fallback:
            chosen = dataset_fallback[task_id]
        if chosen:
            resolved[task_id] = chosen
    return resolved


def parse_offsets_csv(text: str) -> list[int]:
    vals: list[int] = []
    for part in str(text).split(","):
        p = part.strip()
        if not p:
            continue
        vals.append(int(p))
    if not vals:
        raise ValueError("offset list is empty")
    if 0 not in vals:
        raise ValueError("offset list must include 0")
    return vals


def stable_item_key(item: dict[str, Any]) -> str:
    return "|".join(
        [
            str(item.get("task_type", "")),
            str(item.get("task_id", "")),
            str(item.get("episode_id", item.get("episode_index", ""))),
            str(item.get("frame_index", "")),
            str(item.get("camera", "")),
        ]
    )


def shuffled_t3_item(item: dict[str, Any], seed: int) -> dict[str, Any]:
    if str(item.get("task_type", "")) != "T3":
        return item
    choices = item.get("choices", {})
    if not isinstance(choices, dict) or set(choices.keys()) != {"A", "B", "C", "D"}:
        return item
    gt = str(item.get("answer", ""))
    if gt not in {"A", "B", "C", "D"}:
        return item

    x = dict(item)
    labels = ["A", "B", "C", "D"]
    vals = [choices[k] for k in labels]
    key = stable_item_key(item)
    h = hashlib.sha256(f"{seed}|{key}".encode("utf-8")).digest()
    local_rng = random.Random(int.from_bytes(h[:8], byteorder="big", signed=False))
    local_rng.shuffle(vals)
    x["choices"] = {k: v for k, v in zip(labels, vals)}
    x["answer"] = labels[vals.index(choices[gt])]
    x["t3_choice_shuffle_seed"] = int(seed)
    x["t3_choice_shuffle_key"] = key
    return x


def normalize_item_for_eval(item: dict[str, Any]) -> dict[str, Any]:
    """Fill task-specific fields so all items can be evaluated uniformly."""
    x = dict(item)
    task_type = str(x.get("task_type", ""))

    if task_type == "T4":
        if not str(x.get("question", "")).strip():
            x["question"] = (
                "Across these time-ordered frames, which bimanual activity state best describes the robot?"
            )
        if not isinstance(x.get("choices"), dict) or not x.get("choices"):
            x["choices"] = dict(T4_CHOICES)
        if not str(x.get("answer", "")).strip() and "label_id" in x:
            x["answer"] = T4_LABEL_ID_TO_ANSWER.get(int(x["label_id"]), "INVALID")
    return x


def get_frame_paths(item: dict[str, Any], frame_dirs: dict[str, Path], t3_offsets: list[int]) -> list[Path]:
    task_type = str(item["task_type"])
    task_id = str(item["task_id"])
    ep = int(item.get("episode_id", item.get("episode_index", -1)))
    camera = str(item.get("camera", "camera_top"))
    dataset = normalize_text(item.get("dataset", "")).upper()
    recording_id = str(item.get("recording_id", task_id))

    def single(frame_index: int, suffix: str = "", base_key: str = "default") -> Path:
        base = frame_dirs[base_key]
        if _uses_recording_id_frame_names(dataset):
            candidates = _candidate_reassemble_single_paths(base, recording_id, int(frame_index), camera)
        else:
            candidates = _candidate_single_paths(base, task_id, ep, int(frame_index), camera, suffix=suffix)
        return _pick_first_existing(candidates)

    if _uses_recording_id_frame_names(dataset):
        if isinstance(item.get("frame_indices"), list) and item.get("frame_indices"):
            return [single(int(fi)) for fi in item["frame_indices"]]
        if "frame_index" in item:
            return [single(int(item["frame_index"]))]
        raise KeyError(f"{dataset} item missing both frame_index and frame_indices")

    if task_type == "T3":
        # Prefer explicit multi-frame indices when the benchmark item already
        # carries them. This is the current GM100/AIST/RH20T canonical format.
        if isinstance(item.get("frame_indices"), list) and item.get("frame_indices"):
            return [single(int(fi)) for fi in item["frame_indices"]]

        # Backward-compatible fallback for older T3 frame dumps that encoded
        # offsets through suffixed filenames around a center frame.
        fi = int(item["frame_index"])
        out: list[Path] = []
        for off in t3_offsets:
            suf = "t0" if off == 0 else f"t{off:+d}"
            cands = _candidate_multi_paths(
                frame_dirs["T3_multi"],
                task_id=task_id,
                episode_id=ep,
                center_frame_index=fi,
                camera=camera,
                offset=off,
                suffix=suf,
            )
            out.append(_pick_first_existing(cands))
        return out

    if task_type == "T4" and isinstance(item.get("frame_indices"), list):
        return [single(int(fi)) for fi in item["frame_indices"]]

    if task_type == "T6":
        fi = int(item["frame_index"])
        t6_offsets = [(-6, "t-6"), (-3, "t-3"), (0, "t0"), (3, "t+3"), (6, "t+6")]
        out: list[Path] = []
        for off, suf in t6_offsets:
            cands = _candidate_multi_paths(
                frame_dirs["T6_multi"],
                task_id=task_id,
                episode_id=ep,
                center_frame_index=fi,
                camera=camera,
                offset=off,
                suffix=suf,
            )
            out.append(_pick_first_existing(cands))
        return out

    if task_type == "T_progress" and isinstance(item.get("frame_indices"), list):
        return [single(int(fi)) for fi in item["frame_indices"]]

    if task_type == "T_temporal":
        return [single(int(fi)) for fi in item["frame_indices"]]

    if task_type == "T9":
        return [single(int(fi)) for fi in item["frame_indices"]]

    if task_type == "T_binary":
        return [single(int(fi)) for fi in item["frame_indices"]]

    if task_type == "T2":
        # Backward compatible:
        # - legacy T2: pairwise frame_A/frame_B
        # - new T2_contact: single frame via episode_id/frame_index
        if "frame_A" in item and "frame_B" in item:
            fa = item["frame_A"]
            fb = item["frame_B"]
            pa = _pick_first_existing(
                _candidate_single_paths(
                    frame_dirs["default"],
                    task_id,
                    int(fa["episode_id"]),
                    int(fa["frame_index"]),
                    camera,
                )
            )
            pb = _pick_first_existing(
                _candidate_single_paths(
                    frame_dirs["default"],
                    task_id,
                    int(fb["episode_id"]),
                    int(fb["frame_index"]),
                    camera,
                )
            )
            return [pa, pb]
        if "episode_id" in item and "frame_index" in item:
            return [single(int(item["frame_index"]))]
        raise KeyError("T2 item missing both pair fields (frame_A/frame_B) and single-frame fields (episode_id/frame_index)")

    fi = int(item["frame_index"])
    return [single(fi)]


def _answer_protocol_text(item: dict[str, Any]) -> str:
    task_type = str(item.get("task_type", ""))
    if task_type == "T_temporal":
        labels = [str(x).upper() for x in item.get("shuffled_labels", ["X", "Y", "Z"])]
        example = "".join(labels)
        return (
            "Output protocol:\n"
            "- You may include brief reasoning before the final answer.\n"
            f"- The final line must be exactly: <ANSWER>{example}</ANSWER>\n"
            "- Do not output anything after </ANSWER>."
        )

    if task_type == "T_binary":
        choice_keys = {str(k).upper() for k in item.get("choices", {}).keys()}
        if choice_keys == {"YES", "NO"}:
            example = "YES"
        else:
            labels = sorted(set(_t_binary_display_labels(item, num_panels=2)))
            example = labels[0] if labels else "X"
        return (
            "Output protocol:\n"
            "- You may include brief reasoning before the final answer.\n"
            f"- The final line must be exactly: <ANSWER>{example}</ANSWER>\n"
            "- Do not output anything after </ANSWER>."
        )

    choices = item.get("choices", {})
    if isinstance(choices, dict) and choices:
        valid = [str(k).upper() for k in choices.keys()]
    elif task_type == "T4":
        valid = ["A", "B", "C", "D"]
    else:
        valid = []
    example = valid[0] if valid else "A"
    return (
        "Output protocol:\n"
        "- You may include brief reasoning before the final answer.\n"
        f"- The final line must be exactly: <ANSWER>{example}</ANSWER>\n"
        "- Do not output anything after </ANSWER>."
    )


def format_prompt(item: dict[str, Any]) -> str:
    task_type = str(item["task_type"])
    question = str(item.get("question", "")).strip()
    choices = item.get("choices", {})
    answer_protocol = _answer_protocol_text(item)

    prompt_body = ""
    if task_type == "T_binary":
        labels = sorted(set(_t_binary_display_labels(item, num_panels=2)))
        first = labels[0] if len(labels) >= 1 else "X"
        second = labels[1] if len(labels) >= 2 else "Y"
        choice_keys = {str(k).upper() for k in item.get("choices", {}).keys()}
        if choice_keys == {"YES", "NO"}:
            prompt_body = (
                "A single comparison image shows two labeled robot-manipulation panels from the same episode.\n"
                "The left-right placement of the panels and the labels are arbitrary identifiers and do not indicate temporal order.\n"
                f"Is the following statement true: Image {first} happened earlier than Image {second}?\n"
                "Choose exactly one of YES or NO."
            )
        else:
            prompt_body = (
                "A single comparison image shows two labeled robot-manipulation panels from the same episode.\n"
                "The left-right placement of the panels and the labels are arbitrary identifiers and do not indicate temporal order.\n"
                f"Which labeled panel happened earlier in the real manipulation sequence, {first} or {second}?\n"
                f"Choose exactly one label: {first} or {second}."
            )

    elif task_type == "T_temporal":
        labels = [str(x).upper() for x in item.get("shuffled_labels", ["X", "Y", "Z"])]
        ex = f"{labels[1]}{labels[0]}{labels[2]}"
        prompt_body = (
            "You are shown 3 frames from a robot manipulation task.\n"
            f"The frames are labeled {labels[0]}, {labels[1]}, {labels[2]} "
            "(these labels are arbitrary identifiers, not positional, and not time-ordered).\n"
            "Determine the correct chronological order of these frames "
            "(from earliest to latest).\n"
            f"Choose exactly one 3-letter permutation, for example: {ex} means "
            f"{labels[1]} happened first, then {labels[0]}, then {labels[2]}."
        )

    elif isinstance(choices, dict) and choices:
        choices_text = "\n".join(f"{k}: {v}" for k, v in choices.items())
        prompt_body = f"{question}\n\n{choices_text}\n\nChoose exactly one option label."

    else:
        prompt_body = f"{question}\n\nChoose exactly one answer."

    prompt_body = f"{prompt_body}\n\n{answer_protocol}"

    task_meta = normalize_text(item.get("task_meta_description", ""))
    if not task_meta:
        return prompt_body

    return (
        "You are given image(s) from a robot manipulation episode.\n\n"
        f'Task context: The overall task in this episode is "{task_meta}".\n\n'
        "This task context is provided only as background. Do not rely on the task name alone. "
        "Answer based on the visual evidence in the provided image(s).\n\n"
        "Now answer the following question:\n"
        f"{prompt_body}"
    )

def parse_answer(raw: str, item: dict[str, Any]) -> str:
    task_type = str(item["task_type"])
    raw_u = raw.strip().upper().replace("。", ".")

    def _extract_tag_payload(text: str) -> str:
        matches = list(ANSWER_TAG_RE.finditer(text))
        if not matches:
            return ""
        return str(matches[-1].group(1)).strip()

    def _parse_from_valid_set(text: str, valid_set: set[str]) -> str:
        if not valid_set:
            return "INVALID"
        text_u = str(text).strip().upper().replace("。", ".")
        alt = "|".join(sorted((re.escape(v) for v in valid_set), key=len, reverse=True))

        text_compact = re.sub(r"[\s\.\:\-\(\)\[\]\{\}<>/]+", "", text_u)
        if text_compact in valid_set:
            return text_compact

        cue_pat = rf"(?:FINAL\s+ANSWER|ANSWER|OPTION|CHOICE|FINAL)\s*(?:IS|:|=)?\s*\(?({alt})\)?(?![A-Z0-9])"
        cue_matches = list(re.finditer(cue_pat, text_u))
        if cue_matches:
            return cue_matches[-1].group(1)

        tok_pat = rf"(?<![A-Z0-9])({alt})(?![A-Z0-9])"
        tok_matches = list(re.finditer(tok_pat, text_u))
        if tok_matches:
            return tok_matches[-1].group(1)

        return "INVALID"

    def _parse_temporal_text(text: str, labels: list[str]) -> str:
        text_u = str(text).strip().upper().replace("。", ".")
        label_set = set(labels)
        perms = {"".join(p) for p in itertools.permutations(labels, 3)}
        chars = "".join(sorted(label_set))

        exact_hits = []
        for m in re.finditer(rf"(?<![A-Z])([{chars}]{{3}})(?![A-Z])", text_u):
            cand = m.group(1)
            if cand in perms:
                exact_hits.append(cand)
        if exact_hits:
            return exact_hits[-1]

        sep_hits = []
        sep_pat = rf"(?<![A-Z])([{chars}])(?:\s*[,>\-\/|]+\s*|\s+)([{chars}])(?:\s*[,>\-\/|]+\s*|\s+)([{chars}])(?![A-Z])"
        for m in re.finditer(sep_pat, text_u):
            cand = "".join(m.groups())
            if cand in perms:
                sep_hits.append(cand)
        if sep_hits:
            return sep_hits[-1]

        standalone = re.findall(rf"(?<![A-Z])([{chars}])(?![A-Z])", text_u)
        if len(standalone) >= 3:
            window_hits = []
            for i in range(len(standalone) - 2):
                cand = "".join(standalone[i : i + 3])
                if cand in perms:
                    window_hits.append(cand)
            if window_hits:
                return window_hits[-1]

        filtered = "".join(ch for ch in text_u if ch in label_set)
        if len(filtered) >= 3:
            cand = filtered[-3:]
            if cand in perms:
                return cand
            cand = filtered[:3]
            if cand in perms:
                return cand
        return "INVALID"

    if task_type == "T_binary":
        choice_keys = {str(k).upper() for k in item.get("choices", {}).keys()}
        if choice_keys == {"YES", "NO"}:
            valid_set = {"YES", "NO"}
            tagged = _extract_tag_payload(raw)
            if tagged:
                tagged_pred = _parse_from_valid_set(tagged, valid_set)
                if tagged_pred != "INVALID":
                    return tagged_pred
            return _parse_from_valid_set(raw_u, valid_set)

        label_to_answer = _t_binary_label_to_answer(item)
        valid_set = set(label_to_answer)
        tagged = _extract_tag_payload(raw)
        if tagged:
            tagged_pred = _parse_from_valid_set(tagged, valid_set)
            if tagged_pred != "INVALID":
                return label_to_answer[tagged_pred]
        pred = _parse_from_valid_set(raw_u, valid_set)
        if pred != "INVALID":
            return label_to_answer[pred]
        return "INVALID"

    if task_type == "T_temporal":
        labels = [str(x).upper() for x in item.get("shuffled_labels", ["X", "Y", "Z"])]
        tagged = _extract_tag_payload(raw)
        if tagged:
            tagged_pred = _parse_temporal_text(tagged, labels)
            if tagged_pred != "INVALID":
                return tagged_pred
        return _parse_temporal_text(raw_u, labels)

    valid = [str(k).upper() for k in item.get("choices", {}).keys()]
    if not valid and task_type == "T4":
        valid = ["A", "B", "C", "D"]
    if not valid:
        return "INVALID"

    valid_set = set(valid)
    tagged = _extract_tag_payload(raw)
    if tagged:
        tagged_pred = _parse_from_valid_set(tagged, valid_set)
        if tagged_pred != "INVALID":
            return tagged_pred

    pred = _parse_from_valid_set(raw_u, valid_set)
    if pred != "INVALID":
        return pred

    # Last fallback: map textual choice content back to its label.
    choices = item.get("choices", {})
    if isinstance(choices, dict):
        raw_norm = re.sub(r"\s+", " ", raw_u).strip()
        hits: list[tuple[int, str]] = []
        for k, v in choices.items():
            kk = str(k).upper()
            if kk not in valid_set:
                continue
            vv = re.sub(r"\s+", " ", str(v).upper()).strip()
            if not vv:
                continue
            pos = raw_norm.find(vv)
            if pos >= 0:
                hits.append((pos, kk))
        if hits:
            hits.sort(key=lambda x: x[0])
            return hits[-1][1]

    return "INVALID"

def build_request_content(item: dict[str, Any], frame_paths: list[Path]) -> list[dict[str, Any]]:
    content: list[dict[str, Any]] = []
    if str(item.get("task_type", "")) == "T_binary":
        composite_b64 = build_t_binary_composite_b64(frame_paths, item)
        content.append(
            {
                "type": "image_url",
                "image_url": {"url": f"data:image/jpeg;base64,{composite_b64}"},
            }
        )
    else:
        for p in frame_paths:
            content.append(
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:image/jpeg;base64,{img_to_b64(p)}"},
                }
            )
    content.append({"type": "text", "text": format_prompt(item)})
    return content


def assistant_message_text(message: dict[str, Any]) -> str:
    content = message.get("content", "")
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts: list[str] = []
        for chunk in content:
            if isinstance(chunk, dict) and chunk.get("type") == "text":
                parts.append(str(chunk.get("text", "")))
            elif isinstance(chunk, str):
                parts.append(chunk)
        return "".join(parts)
    return str(content or "")


def assistant_reasoning_text(message: dict[str, Any]) -> str:
    reasoning = message.get("reasoning_content", "")
    if isinstance(reasoning, str):
        return reasoning
    if isinstance(reasoning, list):
        return "".join(str(x) for x in reasoning)
    return str(reasoning or "")


def effective_completion_max_tokens(model: str, requested: int) -> int:
    model_l = str(model).lower().strip()
    if model_l.startswith("gpt-5") and requested < 256:
        return 256
    if model_l.startswith("glm-") and requested < 64:
        return 64
    return requested


async def call_one(
    session: aiohttp.ClientSession,
    sem: asyncio.Semaphore,
    item: dict[str, Any],
    frame_dirs: dict[str, Path],
    api_base: str,
    api_key: str,
    model: str,
    max_tokens: int,
    temperature: float,
    retries: int,
    timeout_sec: float,
    t3_offsets: list[int],
) -> dict[str, Any]:
    x = normalize_item_for_eval(item)
    completion_max_tokens = effective_completion_max_tokens(model, max_tokens)

    async with sem:
        try:
            frame_paths = get_frame_paths(x, frame_dirs=frame_dirs, t3_offsets=t3_offsets)
        except Exception as e:  # noqa: BLE001
            return {
                "task_type": x.get("task_type", ""),
                "task_id": x.get("task_id", ""),
                "arm_type": x.get("arm_type", ""),
                "gt_answer": x.get("answer", ""),
                "vlm_answer": "ERROR",
                "correct": False,
                "model": model,
                "raw_response": "",
                "error": f"frame_path_build_error: {e}",
            }
        frame_paths = resolve_missing_frame_paths(x, frame_paths, t3_offsets=t3_offsets)
        missing = [str(p) for p in frame_paths if not p.exists()]
        if missing:
            return {
                "task_type": x.get("task_type", ""),
                "task_id": x.get("task_id", ""),
                "arm_type": x.get("arm_type", ""),
                "gt_answer": x.get("answer", ""),
                "vlm_answer": "MISSING_FRAME",
                "correct": False,
                "model": model,
                "raw_response": "",
                "error": f"missing frames: {missing[:3]}",
            }

        try:
            content = build_request_content(x, frame_paths)
        except Exception as e:  # noqa: BLE001
            return {
                "task_type": x.get("task_type", ""),
                "task_id": x.get("task_id", ""),
                "arm_type": x.get("arm_type", ""),
                "gt_answer": x.get("answer", ""),
                "vlm_answer": "ERROR",
                "correct": False,
                "model": model,
                "raw_response": "",
                "error": f"frame_render_error: {e}",
            }

        payload = {
            "model": model,
            "messages": [{"role": "user", "content": content}],
            "max_tokens": completion_max_tokens,
            "temperature": temperature,
        }

        err = ""
        raw = ""
        for attempt in range(retries):
            try:
                async with session.post(
                    f"{api_base.rstrip('/')}/chat/completions",
                    headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                    json=payload,
                    timeout=aiohttp.ClientTimeout(total=timeout_sec),
                ) as resp:
                    data = await resp.json()
                    if resp.status >= 400:
                        raise RuntimeError(f"HTTP {resp.status}: {data}")
                    message = data["choices"][0]["message"]
                    raw = assistant_message_text(message)
                    reasoning = assistant_reasoning_text(message)
                    if not str(raw).strip() and str(reasoning).strip():
                        raise RuntimeError(
                            "empty_content_with_reasoning_only: "
                            f"finish_reason={data['choices'][0].get('finish_reason', '')}, "
                            f"reasoning_prefix={reasoning[:160]!r}"
                        )
                    pred = parse_answer(raw, x)
                    gt = str(x.get("answer", ""))
                    return {
                        "task_type": x.get("task_type", ""),
                        "task_id": x.get("task_id", ""),
                        "arm_type": x.get("arm_type", ""),
                        "gt_answer": gt,
                        "vlm_answer": pred,
                        "correct": bool(pred == gt),
                        "model": model,
                        "raw_response": raw,
                        "error": "",
                    }
            except Exception as e:  # noqa: BLE001
                err = str(e)
                if attempt < retries - 1:
                    await asyncio.sleep(2**attempt)

        return {
            "task_type": x.get("task_type", ""),
            "task_id": x.get("task_id", ""),
            "arm_type": x.get("arm_type", ""),
            "gt_answer": x.get("answer", ""),
            "vlm_answer": "ERROR",
            "correct": False,
            "model": model,
            "raw_response": raw,
            "error": err,
        }



def preferred_api_key_envs(model: str, api_base: str) -> list[str]:
    model_l = str(model).lower()
    base_l = str(api_base).lower()
    if "claude" in model_l or "anthropic" in base_l:
        return ["ANTHROPIC_API_KEY", "OPENAI_API_KEY", "DASHSCOPE_API_KEY"]
    if "qwen" in model_l or "dashscope" in base_l:
        return ["DASHSCOPE_API_KEY", "OPENAI_API_KEY"]
    return ["OPENAI_API_KEY", "DASHSCOPE_API_KEY", "ANTHROPIC_API_KEY"]


async def auth_preflight(
    api_base: str,
    api_key: str,
    model: str,
    timeout_sec: float,
    max_tokens: int,
    retries: int,
) -> None:
    url = f"{api_base.rstrip('/')}/chat/completions"
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": [{"type": "text", "text": "Reply with A."}]}],
        "max_tokens": effective_completion_max_tokens(model, max_tokens),
        "temperature": 0.0,
    }

    status = 0
    body: dict[str, Any] = {}
    async with aiohttp.ClientSession() as session:
        for attempt in range(retries):
            try:
                async with session.post(
                    url,
                    headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                    json=payload,
                    timeout=aiohttp.ClientTimeout(total=timeout_sec),
                ) as resp:
                    status = int(resp.status)
                    try:
                        body = await resp.json()
                    except Exception:  # noqa: BLE001
                        body = {"raw": await resp.text()}
                if status < 500:
                    break
            except Exception as e:  # noqa: BLE001
                body = {"raw": str(e)}
            if attempt < retries - 1:
                await asyncio.sleep(2**attempt)

    if status in {401, 403}:
        raise RuntimeError(
            f"Authentication failed in preflight (HTTP {status}). "
            f"Check --api-key / provider env. Response: {body}"
        )
    if status >= 500:
        raise RuntimeError(f"Server error in preflight (HTTP {status}): {body}")
    if status >= 400:
        print(f"[WARN] Preflight returned HTTP {status}; continue anyway. Body: {body}")


async def vision_preflight(
    api_base: str,
    api_key: str,
    model: str,
    timeout_sec: float,
    max_tokens: int,
    retries: int,
) -> None:
    sample_path = next(
        iter(sorted((BENCHMARK_DIR / "benchmark_v1_frames_tbinary_20260330").glob("*.jpg"))),
        None,
    )
    if sample_path is None:
        raise RuntimeError("Vision preflight could not find any local sample image.")

    effective_tokens = effective_completion_max_tokens(model, max_tokens)
    payload = {
        "model": model,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/jpeg;base64,{img_to_b64(sample_path)}"},
                    },
                    {"type": "text", "text": "Reply with only A."},
                ],
            }
        ],
        "max_tokens": effective_tokens,
        "temperature": 0.0,
    }
    url = f"{api_base.rstrip('/')}/chat/completions"

    last_err = ""
    async with aiohttp.ClientSession() as session:
        for attempt in range(retries):
            try:
                async with session.post(
                    url,
                    headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                    json=payload,
                    timeout=aiohttp.ClientTimeout(total=timeout_sec),
                ) as resp:
                    status = int(resp.status)
                    try:
                        body = await resp.json()
                    except Exception:  # noqa: BLE001
                        body = {"raw": await resp.text()}

                if status >= 400:
                    last_err = (
                        "Vision preflight failed. "
                        f"Model={model}, HTTP {status}, body={body}. "
                        "This usually means the selected model or route does not accept image inputs on the current gateway."
                    )
                    if status < 500:
                        raise RuntimeError(last_err)
                else:
                    try:
                        message = body["choices"][0]["message"]
                    except Exception as e:  # noqa: BLE001
                        raise RuntimeError(f"Vision preflight returned an unexpected response shape: {body}") from e

                    content = assistant_message_text(message).strip()
                    reasoning = assistant_reasoning_text(message).strip()
                    finish_reason = str(body["choices"][0].get("finish_reason", "")).strip()
                    if content:
                        return
                    last_err = (
                        "Vision preflight produced no final answer content. "
                        f"Model={model}, finish_reason={finish_reason}, "
                        f"reasoning_prefix={reasoning[:160]!r}. "
                        "This model is not compatible with the current benchmark protocol on this gateway."
                    )
            except Exception as e:  # noqa: BLE001
                last_err = str(e)
                if "HTTP 4" in last_err:
                    raise
            if attempt < retries - 1:
                await asyncio.sleep(2**attempt)
    raise RuntimeError(last_err)


async def benchmark_item_preflight(
    api_base: str,
    api_key: str,
    model: str,
    timeout_sec: float,
    max_tokens: int,
    retries: int,
    item: dict[str, Any],
    frame_dirs: dict[str, Path],
    t3_offsets: list[int],
) -> None:
    x = normalize_item_for_eval(item)
    frame_paths = get_frame_paths(x, frame_dirs=frame_dirs, t3_offsets=t3_offsets)
    frame_paths = resolve_missing_frame_paths(x, frame_paths, t3_offsets=t3_offsets)
    missing = [str(p) for p in frame_paths if not p.exists()]
    if missing:
        raise RuntimeError(f"Benchmark preflight missing frames: {missing[:3]}")

    payload = {
        "model": model,
        "messages": [{"role": "user", "content": build_request_content(x, frame_paths)}],
        "max_tokens": effective_completion_max_tokens(model, max_tokens),
        "temperature": 0.0,
    }
    url = f"{api_base.rstrip('/')}/chat/completions"

    last_err = ""
    async with aiohttp.ClientSession() as session:
        for attempt in range(retries):
            try:
                async with session.post(
                    url,
                    headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                    json=payload,
                    timeout=aiohttp.ClientTimeout(total=timeout_sec),
                ) as resp:
                    status = int(resp.status)
                    try:
                        body = await resp.json()
                    except Exception:  # noqa: BLE001
                        body = {"raw": await resp.text()}

                if status >= 400:
                    last_err = (
                        "Benchmark-item preflight failed. "
                        f"Model={model}, HTTP {status}, body={body}."
                    )
                    if status < 500:
                        raise RuntimeError(last_err)
                else:
                    try:
                        message = body["choices"][0]["message"]
                    except Exception as e:  # noqa: BLE001
                        raise RuntimeError(f"Benchmark-item preflight returned an unexpected response shape: {body}") from e

                    content = assistant_message_text(message).strip()
                    reasoning = assistant_reasoning_text(message).strip()
                    finish_reason = str(body["choices"][0].get("finish_reason", "")).strip()
                    if content:
                        return
                    last_err = (
                        "Benchmark-item preflight produced no final answer content. "
                        f"Model={model}, finish_reason={finish_reason}, "
                        f"reasoning_prefix={reasoning[:200]!r}. "
                        "This model is not suitable for the current benchmark protocol on this gateway."
                    )
            except Exception as e:  # noqa: BLE001
                last_err = str(e)
                if "HTTP 4" in last_err:
                    raise
            if attempt < retries - 1:
                await asyncio.sleep(2**attempt)
    raise RuntimeError(last_err)

async def run_eval(args: argparse.Namespace) -> None:
    api_key = str(args.api_key).strip()
    env_candidates = preferred_api_key_envs(args.model, args.api_base)
    if not api_key:
        for env_key in env_candidates:
            v = os.getenv(env_key, "").strip()
            if v:
                api_key = v
                break
    if not api_key:
        api_key = DEFAULT_API_KEY.strip()
    if not api_key:
        raise RuntimeError(
            "Missing API key. Set --api-key or env "
            f"{'/'.join(env_candidates)}."
        )

    await auth_preflight(
        api_base=args.api_base,
        api_key=api_key,
        model=args.model,
        timeout_sec=args.timeout_sec,
        max_tokens=max(32, int(args.max_tokens)),
        retries=max(1, int(args.retries)),
    )
    await vision_preflight(
        api_base=args.api_base,
        api_key=api_key,
        model=args.model,
        timeout_sec=args.timeout_sec,
        max_tokens=max(64, int(args.max_tokens)),
        retries=max(1, int(args.retries)),
    )

    frame_dirs = {
        "default": Path(args.frame_dir_default),
        "T3_multi": Path(args.frame_dir_t3_multi),
        "T6_multi": Path(args.frame_dir_t6_multi),
    }
    for k, v in frame_dirs.items():
        if not v.exists():
            raise FileNotFoundError(f"Frame dir missing ({k}): {v}")

    items: list[dict[str, Any]] = []
    task_meta_map: dict[str, str] = {}
    if args.prepend_task_meta:
        task_meta_map = load_task_meta_descriptions(Path(args.task_meta_xlsx).resolve())

    with open(args.input, "r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                item = json.loads(line)
                if args.prepend_task_meta:
                    task_id = str(item.get("task_id", "")).strip()
                    item["task_meta_description"] = task_meta_map.get(task_id, "")
                items.append(item)

    if args.limit > 0:
        items = items[: args.limit]
    t3_offsets = parse_offsets_csv(args.t3_offsets)
    if args.t3_randomize_choices:
        items = [shuffled_t3_item(x, seed=args.t3_randomize_seed) for x in items]

    if items:
        await benchmark_item_preflight(
            api_base=args.api_base,
            api_key=api_key,
            model=args.model,
            timeout_sec=args.timeout_sec,
            max_tokens=int(args.max_tokens),
            retries=max(1, int(args.retries)),
            item=items[0],
            frame_dirs=frame_dirs,
            t3_offsets=t3_offsets,
        )

    prewarm_summary = prewarm_gm100_missing_frames(
        items=items,
        frame_dirs=frame_dirs,
        t3_offsets=t3_offsets,
    )

    sem = asyncio.Semaphore(args.concurrency)
    start = time.time()

    async with aiohttp.ClientSession() as session:
        tasks = [
            call_one(
                session=session,
                sem=sem,
                item=item,
                frame_dirs=frame_dirs,
                api_base=args.api_base,
                api_key=api_key,
                model=args.model,
                max_tokens=args.max_tokens,
                temperature=args.temperature,
                retries=args.retries,
                timeout_sec=args.timeout_sec,
                t3_offsets=t3_offsets,
            )
            for item in items
        ]
        results = await tqdm_asyncio.gather(*tasks)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        for r in results:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    valid = [r for r in results if r["vlm_answer"] not in {"MISSING_FRAME", "ERROR", "INVALID"}]
    acc = (sum(1 for r in valid if r["correct"]) / len(valid)) if valid else 0.0
    task_meta_attached_items = sum(1 for x in items if normalize_text(x.get("task_meta_description", "")))
    task_meta_missing_task_ids = sorted(
        {str(x.get("task_id", "")).strip() for x in items if not normalize_text(x.get("task_meta_description", ""))}
    )
    summary = {
        "input": args.input,
        "output": str(out_path),
        "model": args.model,
        "num_items": len(items),
        "num_valid": len(valid),
        "overall_acc_valid_only": acc,
        "prepend_task_meta": bool(args.prepend_task_meta),
        "task_meta_xlsx": str(Path(args.task_meta_xlsx).resolve()) if args.prepend_task_meta else "",
        "task_meta_attached_items": task_meta_attached_items,
        "task_meta_missing_task_ids": task_meta_missing_task_ids if args.prepend_task_meta else [],
        "t3_offsets": t3_offsets,
        "t3_randomize_choices": bool(args.t3_randomize_choices),
        "t3_randomize_seed": int(args.t3_randomize_seed),
        "requested_max_tokens": int(args.max_tokens),
        "effective_max_tokens": int(effective_completion_max_tokens(args.model, args.max_tokens)),
        "gm100_prewarm": prewarm_summary,
        "elapsed_sec": time.time() - start,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser()
    p.add_argument("--input", default=DEFAULT_INPUT_JSONL)
    p.add_argument("--output", default=DEFAULT_OUTPUT_JSONL)

    p.add_argument("--api-base", default=DEFAULT_API_BASE)
    p.add_argument("--api-key", default="")
    p.add_argument("--model", default=DEFAULT_MODEL)
    p.add_argument(
        "--prepend-task-meta",
        action="store_true",
        help="Prepend task-level meta description to each prompt as episode background.",
    )
    p.add_argument(
        "--task-meta-xlsx",
        default=DEFAULT_TASK_META_XLSX,
        help='Workbook used to resolve task meta descriptions (expects sheet "link").',
    )

    p.add_argument("--concurrency", type=int, default=DEFAULT_CONCURRENCY)
    p.add_argument("--max-tokens", type=int, default=DEFAULT_MAX_TOKENS)
    p.add_argument("--temperature", type=float, default=DEFAULT_TEMPERATURE)
    p.add_argument("--retries", type=int, default=DEFAULT_RETRIES)
    p.add_argument("--timeout-sec", type=float, default=DEFAULT_TIMEOUT_SEC)
    p.add_argument("--limit", type=int, default=DEFAULT_LIMIT, help="0 means all items")

    p.add_argument(
        "--frame-dir-default",
        default=DEFAULT_FRAME_DIR,
    )
    p.add_argument(
        "--frame-dir-t3-multi",
        default=DEFAULT_FRAME_DIR_T3,
    )
    p.add_argument(
        "--t3-offsets",
        default=DEFAULT_T3_OFFSETS,
        help="Comma-separated T3 offsets, default: -10,-5,0,+5",
    )
    p.add_argument(
        "--t3-randomize-choices",
        action="store_true",
        help="For T3 only, deterministically reshuffle A/B/C/D semantics per item as bias control.",
    )
    p.add_argument(
        "--t3-randomize-seed",
        type=int,
        default=20260320,
        help="Seed used for deterministic per-item T3 choice reshuffling.",
    )
    p.add_argument(
        "--frame-dir-t6-multi",
        default=DEFAULT_FRAME_DIR_T6,
    )
    return p


if __name__ == "__main__":
    args = build_argparser().parse_args()
    asyncio.run(run_eval(args))
